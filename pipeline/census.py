"""Report every distinct header layout in the downloaded archive.

The era maps in pipeline/mappings/ have to cover every header any source has
ever shipped, and extraction fails loudly on anything unmapped. This tool is
how those maps get authored and maintained: it opens every file the manifests
pin, reads only the header row, and groups periods by identical layout.

Run it after acquiring new periods. If it shows a layout the era map does not
cover, that is a decision to make, not a bug to route around.

Usage: python pipeline/census.py [--system SYSTEM] [--json]
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
import zipfile
from pathlib import Path

import common

SNIFF = 1 << 16


def _header_from_csv_bytes(raw: bytes) -> list[str]:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        return ["<undecodable>"]
    line = text.splitlines()[0] if text.splitlines() else ""
    return next(csv.reader(io.StringIO(line)), [])


def _header_from_xlsx(path: Path, member: str | None = None) -> list[str]:
    import openpyxl

    source: Path | io.BytesIO = path
    if member is not None:
        with zipfile.ZipFile(path) as zf:
            source = io.BytesIO(zf.read(member))
    book = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            return ["" if c is None else str(c) for c in row]
        return []
    finally:
        book.close()


def headers_in(path: Path) -> list[tuple[str, list[str]]]:
    """Return (label, header) for every tabular member of a file."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("rb") as fh:
            return [(path.name, _header_from_csv_bytes(fh.read(SNIFF)))]
    if suffix == ".xlsx":
        return [(path.name, _header_from_xlsx(path))]
    if suffix == ".zip":
        out: list[tuple[str, list[str]]] = []
        with zipfile.ZipFile(path) as zf:
            for info in zf.infolist():
                name = info.filename
                if name.endswith("/") or name.startswith("__MACOSX"):
                    continue
                low = name.lower()
                if low.endswith(".csv"):
                    with zf.open(info) as fh:
                        out.append((name, _header_from_csv_bytes(fh.read(SNIFF))))
                elif low.endswith(".xlsx"):
                    out.append((name, _header_from_xlsx(path, name)))
        return out
    return []


def census(system_id: str) -> dict[str, list[str]]:
    """Map 'header tuple' -> the members that use it."""
    manifest = common.load_manifest(system_id)
    layouts: dict[tuple[str, ...], list[str]] = {}
    for period in sorted(manifest.get("sources", {})):
        entry = manifest["sources"][period]
        if not entry.get("sha256"):
            continue  # not downloaded yet
        path = common.local_path(system_id, period, entry.get("content_format"))
        if not path.exists():
            continue
        try:
            for label, header in headers_in(path):
                key = tuple(header)
                layouts.setdefault(key, []).append(f"{period}:{label}")
        except (zipfile.BadZipFile, OSError, ValueError) as exc:
            print(f"  {period}: UNREADABLE — {exc}", file=sys.stderr)
    return {" | ".join(k): v for k, v in layouts.items()}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--system", choices=sorted(common.SYSTEMS), action="append")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    result = {}
    for system_id in args.system or sorted(common.SYSTEMS):
        layouts = census(system_id)
        result[system_id] = layouts
        if args.json:
            continue
        print(f"\n=== {system_id} — {len(layouts)} distinct layout(s)")
        for header, members in sorted(layouts.items(), key=lambda kv: -len(kv[1])):
            cols = header.split(" | ")
            print(f"\n  [{len(members)} file(s)] {len(cols)} columns")
            print(f"    {header}")
            sample = ", ".join(members[:4])
            more = f" … +{len(members) - 4}" if len(members) > 4 else ""
            print(f"    e.g. {sample}{more}")
    if args.json:
        print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
