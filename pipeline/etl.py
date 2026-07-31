"""Staged ETL: the acquired archive -> a DuckDB star schema.

  extract  land every file into raw_trips / raw_stations, headers unified by
           the per-system era maps. An unmapped header ABORTS the run.
  clean    type values, parse timestamps, drop the unusable, dedupe
  conform  station identity, canonical month, quality flags
  model    fact_trips + dim_system / dim_station / dim_date / dim_membership

Each stage is re-runnable. Transform logic lives in pipeline/sql/; this module
only orchestrates and records row accounting into etl_metrics.

Usage: python pipeline/etl.py [--stage extract|clean|conform|model|all]
                              [--system SYSTEM] [--limit N]
"""

from __future__ import annotations

import argparse
import io
import json
import shutil
import sys
import zipfile
from pathlib import Path

import duckdb

import common

SQL_DIR = common.PIPELINE_DIR / "sql"
WAREHOUSE = common.DATA_WAREHOUSE / "bikeshare.duckdb"
STAGES = ("extract", "reference", "clean", "conform", "model")
SQL_FILES = {"clean": "20_clean.sql", "conform": "30_conform.sql", "model": "40_model.sql"}

# Sources that exist in the manifest but are deliberately not trip-level.
# Toronto 2014-2015 is a workbook of pre-aggregated OD matrices: no timestamps,
# no per-trip rows. Loading it into fact_trips would mix grains, which is the
# one thing this project's comparability claim cannot survive.
EXCLUDED: dict[tuple[str, str], str] = {
    ("tor-bikeshare", "2014-2015"): "pre-aggregated OD matrices, not trip rows",
}


class UnknownColumns(Exception):
    """A source file has headers the era map does not know about."""


def load_map(system_id: str) -> dict:
    path = common.MAPPINGS_DIR / f"{system_id}.json"
    raw = json.loads(path.read_text(encoding="utf-8"))
    return {
        "trips": {k: v for k, v in raw.get("trips", {}).items() if not k.startswith("_")},
        "stations": {k: v for k, v in raw.get("stations", {}).items() if not k.startswith("_")},
    }


def extracted_dir(system_id: str, period: str) -> Path:
    return common.DATA_RAW / system_id / "_extracted" / period


DATA_SUFFIXES = (".csv", ".xlsx")
# Members that are documentation rather than data. Listed explicitly, because
# the alternative — skipping anything that is not a CSV — is how Toronto's
# November 2022 went missing: that month ships as a zip nested inside the
# annual zip, and a silent skip dropped 320,000 trips from the site.
IGNORABLE_SUFFIXES = (".docx", ".doc", ".pdf", ".txt", ".md", ".rtf", ".xml")


class ReconciliationFailed(Exception):
    """A file landed a different number of records than the source contains."""


class UnknownMember(Exception):
    """An archive member is neither data, a nested archive, nor documentation."""


def _unpack(zf: zipfile.ZipFile, dest_dir: Path, source: str, depth: int = 0) -> None:
    """Extract data members, recursing into nested archives.

    Every member must be classifiable. An unrecognized one ABORTS, exactly as
    an unmapped column header does — the guarantee is about unknown input, and
    an archive member is input.
    """
    if depth > 3:
        raise UnknownMember(f"{source}: archive nested more than 3 deep")
    for info in zf.infolist():
        name = info.filename
        base = Path(name).name
        if name.endswith("/") or name.startswith("__MACOSX") or base.startswith((".", "~$")):
            continue
        low = base.lower()
        if low.endswith(DATA_SUFFIXES):
            dest = dest_dir / base
            if dest.exists():
                # Flattening means two members could collide and one would win
                # silently. Refuse instead.
                raise UnknownMember(
                    f"{source}: two members flatten to the same name {base!r}"
                )
            with zf.open(info) as src, dest.open("wb") as dst:
                shutil.copyfileobj(src, dst, 1 << 20)
        elif low.endswith(".zip"):
            with zf.open(info) as nested_raw:
                payload = io.BytesIO(nested_raw.read())
            with zipfile.ZipFile(payload) as nested:
                _unpack(nested, dest_dir, f"{source} -> {base}", depth + 1)
        elif low.endswith(IGNORABLE_SUFFIXES):
            continue
        else:
            raise UnknownMember(
                f"{source}: member {name!r} is neither data, a nested archive, "
                "nor documentation. Classify it in etl.py — deliberately."
            )


def tabular_files(system_id: str, period: str, entry: dict) -> list[Path]:
    """Every CSV/XLSX for a period, unpacking archives once and caching them.

    The cache is keyed by the source checksum, not just the period: a refreshed
    zip at the same period must invalidate it. BIXI's annual filename encodes
    the months it contains, so a mid-season re-publish changes the URL and the
    pin while the period stays '2026' — a period-keyed cache would happily
    serve January-to-June forever.
    """
    path = common.local_path(system_id, period, entry.get("content_format"))
    if not path.exists():
        raise FileNotFoundError(
            f"{system_id} {period}: pinned in the manifest but absent at {path}. "
            "Run download.py; a missing pinned file must not silently shrink the "
            "warehouse."
        )
    if path.suffix.lower() in DATA_SUFFIXES:
        return [path]

    sha = (entry.get("sha256") or "nopin")[:12]
    out_dir = extracted_dir(system_id, period).with_name(f"{period}.{sha}")
    if not out_dir.exists():
        # Drop any cache built from a different pin for this period.
        for stale in out_dir.parent.glob(f"{period}.*"):
            shutil.rmtree(stale, ignore_errors=True)
        tmp = out_dir.with_suffix(".partial")
        shutil.rmtree(tmp, ignore_errors=True)
        tmp.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(path) as zf:
            _unpack(zf, tmp, f"{system_id} {period}")
        tmp.replace(out_dir)
    return sorted(p for p in out_dir.iterdir() if p.suffix.lower() in DATA_SUFFIXES)


def sheets_in(path: Path) -> list[str]:
    """Every worksheet in a workbook.

    read_xlsx() reads only the FIRST sheet. Toronto's 2016.xlsx carries Q3 and
    Q4 as two sheets, so 217,569 trips — a whole quarter — were never read and
    nothing said so. A workbook is an archive too.
    """
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True)
    try:
        return list(book.sheetnames)
    finally:
        book.close()


SCRUB_DIR_NAME = "_utf8"


def ensure_utf8(path: Path) -> tuple[Path, int]:
    """Return a UTF-8-clean copy of a CSV, plus the number of lines repaired.

    Some Toronto and Vancouver months carry cp1252 bytes in station names — an
    en-dash in "25 York St - Union Station South", a curly apostrophe in
    "King's College Cr.". DuckDB's reader discards those rows, and with
    ignore_errors it does so silently.

    The damage is not proportional to the row count. Toronto station 7070 lost
    eight consecutive months of 2023 and reads as if it closed; 7358 lost its
    first eight months of service. A station with a fabricated history is worse
    than a slightly low total.

    encoding='latin-1' is not a fix — DuckDB 1.5.5 validates C1 bytes and
    rejects these files outright. So the repair happens a line at a time:
    UTF-8 strict, then cp1252, then latin-1 as a last resort.
    """
    if path.suffix.lower() != ".csv":
        return path, 0
    # Stream the check: BIXI's annual CSV is 2.8 GB and must not be slurped.
    import codecs

    decoder = codecs.getincrementaldecoder("utf-8")()
    clean = True
    with path.open("rb") as fh:
        while chunk := fh.read(1 << 22):
            try:
                decoder.decode(chunk)
            except UnicodeDecodeError:
                clean = False
                break
    if clean:
        return path, 0          # already clean, no copy made

    dest = path.parent / SCRUB_DIR_NAME / path.name
    marker = dest.with_suffix(dest.suffix + ".lines")
    if dest.exists() and marker.exists():
        return dest, int(marker.read_text())

    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_suffix(dest.suffix + ".part")
    repaired = 0
    with path.open("rb") as src, tmp.open("wb") as out:
        for line in src:
            try:
                line.decode("utf-8")
            except UnicodeDecodeError:
                repaired += 1
                for enc in ("cp1252", "latin-1"):
                    try:
                        line = line.decode(enc).encode("utf-8")
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    line = line.decode("utf-8", errors="replace").encode("utf-8")
            out.write(line)
    tmp.replace(dest)
    marker.write_text(str(repaired))
    return dest, repaired


def source_record_count(path: Path, sheet: str | None = None) -> int:
    """How many data records the source actually contains.

    A fast line count first; only if that disagrees with what landed does the
    caller pay for a csv.reader pass, which is the one that handles quoted
    embedded newlines correctly.
    """
    if path.suffix.lower() == ".xlsx":
        import openpyxl

        book = openpyxl.load_workbook(path, read_only=True)
        try:
            ws = book[sheet] if sheet else book[book.sheetnames[0]]
            return max(0, ws.max_row - 1)
        finally:
            book.close()
    with path.open("rb") as fh:
        lines = sum(1 for _ in fh)
    return max(0, lines - 1)


def exact_record_count(path: Path) -> int:
    """csv.reader count, for when the line count and the landed count differ."""
    import csv as _csv

    # errors="replace" so a file that is not UTF-8 still yields a COUNT. The
    # point here is to report how many records the source has, which is exactly
    # the situation where the bytes are bad — crashing would replace the gate's
    # explanation with a decode traceback.
    with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
        return max(0, sum(1 for _ in _csv.reader(fh)) - 1)


def reader_expr(path: Path, sheet: str | None = None) -> str:
    quoted = str(path).replace("'", "''")
    if path.suffix.lower() == ".xlsx":
        if sheet is None:
            return f"read_xlsx('{quoted}', all_varchar = true)"
        esc = sheet.replace("'", "''")
        return f"read_xlsx('{quoted}', all_varchar = true, sheet = '{esc}')"
    # ignore_errors keeps one ragged row from killing a 3 GB file.
    #
    # A row the reader dropped here would enter no count — which is why two
    # guards close that door before this expression can lose anything
    # silently: ensure_utf8 repairs invalid-encoding lines (the one measured
    # loss mode, 31,315 lines, station-biased) and records the count per
    # file, and extract_system compares an independently counted
    # source_record_count against rows landed and raises
    # ReconciliationFailed on any mismatch. An earlier version of this
    # comment recorded the loss as an open gap; both halves of the fix are
    # ~100 lines below.
    return (
        f"read_csv('{quoted}', header = true, all_varchar = true, "
        "sample_size = -1, ignore_errors = true, null_padding = true)"
    )


def read_header(con, path: Path, sheet: str | None = None) -> list[str]:
    rows = con.execute(f"DESCRIBE SELECT * FROM {reader_expr(path, sheet)}").fetchall()
    return [r[0] for r in rows]


def tabular_units(system_id: str, period: str, entry: dict) -> list[tuple[Path, str | None, int]]:
    """(file, sheet, lines_repaired) triples to load, CSVs made UTF-8 clean."""
    units: list[tuple[Path, str | None, int]] = []
    for path in tabular_files(system_id, period, entry):
        if path.suffix.lower() == ".xlsx":
            names = sheets_in(path)
            units += ([(path, n, 0) for n in names] if len(names) > 1
                      else [(path, None, 0)])
        else:
            scrubbed, repaired = ensure_utf8(path)
            units.append((scrubbed, None, repaired))
    return units


def plan(header: list[str], column_map: dict, source: str) -> list[tuple[str, str]]:
    """(raw, unified) pairs to load. Raises on any header the map omits."""
    unknown = [h for h in header if h.strip() not in column_map]
    if unknown:
        raise UnknownColumns(
            f"{source}: unmapped header(s) {unknown!r}. Add them to the era map "
            "in pipeline/mappings/ — deliberately, not reflexively."
        )
    pairs = [(h, column_map[h.strip()]) for h in header if column_map[h.strip()]]
    targets = [t for _, t in pairs]
    if len(targets) != len(set(targets)):
        raise UnknownColumns(f"{source}: two headers map to the same column: {targets}")
    return pairs


def classify(header: list[str], maps: dict) -> str | None:
    """Is this file trips, stations, or neither? Decided by header, not name."""
    clean = [h.strip() for h in header]
    if clean and all(h in maps["stations"] for h in clean) and maps["stations"]:
        return "stations"
    if any(h in maps["trips"] for h in clean):
        return "trips"
    return None


def run_extract(con, systems: list[str], limit: int | None) -> None:
    run_sql(con, "10_extract.sql")
    ensure_audit_table(con)
    for system_id in systems:
        con.execute("DELETE FROM raw_file_audit WHERE system_id = ?", [system_id])
    # Reloading a system replaces only that system's rows. Without this, a
    # re-run would double every count instead of refreshing it.
    for system_id in systems:
        con.execute("DELETE FROM raw_trips WHERE system_id = ?", [system_id])
        con.execute("DELETE FROM raw_stations WHERE system_id = ?", [system_id])
    for system_id in systems:
        maps = load_map(system_id)
        manifest = common.load_manifest(system_id)
        loaded = skipped = 0
        for period in sorted(manifest.get("sources", {})):
            entry = manifest["sources"][period]
            if not entry.get("sha256"):
                continue
            if (system_id, period) in EXCLUDED:
                print(f"  {system_id} {period}: excluded — {EXCLUDED[(system_id, period)]}")
                skipped += 1
                continue
            for path, sheet, repaired in tabular_units(system_id, period, entry):
                label = path.name if sheet is None else f"{path.name}#{sheet.strip()}"
                header = read_header(con, path, sheet)
                kind = classify(header, maps)
                if kind is None:
                    raise UnknownColumns(
                        f"{system_id} {period} {label}: header matches neither "
                        f"the trips nor the stations map: {header!r}"
                    )
                pairs = plan(header, maps[kind], f"{system_id} {period} {label}")
                cols = ", ".join(f'"{u}"' for _, u in pairs)
                sel = ", ".join(f'"{r}"' for r, _ in pairs)
                table = "raw_trips" if kind == "trips" else "raw_stations"
                con.execute(
                    f"INSERT INTO {table} (system_id, source_period, source_file, {cols}) "
                    f"SELECT ?, ?, ?, {sel} FROM {reader_expr(path, sheet)}",
                    [system_id, period, label],
                )
                landed = con.execute(
                    f"SELECT count(*) FROM {table} WHERE system_id = ? "
                    "AND source_period = ? AND source_file = ?",
                    [system_id, period, label],
                ).fetchone()[0]
                expected = source_record_count(path, sheet)
                if expected != landed and path.suffix.lower() == ".csv":
                    # Line count and record count differ when a field contains a
                    # quoted newline. Pay for the exact count only when needed.
                    expected = exact_record_count(path)
                con.execute(
                    "INSERT INTO raw_file_audit VALUES (?, ?, ?, ?, ?, ?, ?)",
                    [system_id, period, label, expected, landed, repaired, kind],
                )
                if expected != landed:
                    raise ReconciliationFailed(
                        f"{system_id} {period} {label}: source has {expected:,} "
                        f"records, {landed:,} landed — {expected - landed:,} "
                        "unaccounted for. The funnel begins at rows_landed, so a "
                        "row lost here is invisible to every later count."
                    )
                loaded += 1
                note = f"  (+{repaired:,} lines repaired)" if repaired else ""
                print(f"  {system_id} {period} {label}: {kind}{note}", flush=True)
            if limit and loaded >= limit:
                break
        print(f"{system_id}: {loaded} file(s) loaded, {skipped} excluded")

    record(con, "extract", "rows_landed", "SELECT count(*) FROM raw_trips")
    record(con, "extract", "files_landed",
           "SELECT count(DISTINCT system_id || source_period || source_file) FROM raw_trips")
    record(con, "extract", "station_rows", "SELECT count(*) FROM raw_stations")
    for system_id in systems:
        record(con, "extract", f"rows_{system_id}",
               f"SELECT count(*) FROM raw_trips WHERE system_id = '{system_id}'")
    record_encoding_metrics(con)


def ensure_audit_table(con) -> None:
    """The per-file audit: what the source held, what landed, what was repaired.

    `kind` was added by spec 028 so the quality report can separate trip files
    from station snapshots without a semi-join over 135M rows, and `ALTER ... IF
    NOT EXISTS` is what lets an existing warehouse gain it without a reload.
    """
    con.execute("""CREATE TABLE IF NOT EXISTS raw_file_audit (
        system_id VARCHAR, source_period VARCHAR, source_file VARCHAR,
        source_records BIGINT, rows_landed BIGINT, lines_repaired BIGINT,
        kind VARCHAR)""")
    con.execute("ALTER TABLE raw_file_audit ADD COLUMN IF NOT EXISTS kind VARCHAR")


def record_encoding_metrics(con) -> None:
    """Roll the per-file repair counts up, so the report never hardcodes them.

    The quality report used to state the encoding loss as prose — "~31,000 rows
    discarded on invalid cp1252, an open gap" — written before `ensure_utf8`
    existed and never revised after it did. The rows are not discarded any
    more; they are repaired, a line at a time, and the count is known per file.
    A generated report has no business restating from memory something the
    warehouse can answer.
    """
    # The backfill can run against a warehouse the extract stage has not
    # touched in this session, so it cannot assume 10_extract.sql just ran.
    con.execute("""CREATE TABLE IF NOT EXISTS etl_metrics (
        stage VARCHAR NOT NULL, metric VARCHAR NOT NULL, value BIGINT NOT NULL)""")
    record(con, "extract", "lines_repaired",
           "SELECT coalesce(sum(lines_repaired), 0) FROM raw_file_audit")
    record(con, "extract", "files_encoding_repaired",
           "SELECT count(*) FROM raw_file_audit WHERE lines_repaired > 0")


class MissingRepairMarker(Exception):
    """A scrubbed copy exists but the line count that produced it does not."""


def scrub_marker(path: Path) -> Path:
    """Where ensure_utf8 records how many lines it repaired in `path`."""
    dest = path.parent / SCRUB_DIR_NAME / path.name
    return dest.with_suffix(dest.suffix + ".lines")


def period_dir(system_id: str, period: str, entry: dict) -> Path | None:
    """The directory holding a period's tabular files, WITHOUT unpacking.

    `tabular_files` would rebuild a missing cache and re-read every byte of a
    2.8 GB CSV to decide whether it is UTF-8 clean. The backfill must not do
    that: it is reading what extraction already wrote down.
    """
    path = common.local_path(system_id, period, entry.get("content_format"))
    if path.suffix.lower() in DATA_SUFFIXES:
        return path.parent if path.exists() else None
    sha = (entry.get("sha256") or "nopin")[:12]
    out_dir = extracted_dir(system_id, period).with_name(f"{period}.{sha}")
    return out_dir if out_dir.exists() else None


def backfill_encoding_repairs(con) -> int:
    """Read the repair markers on disk into raw_file_audit, no extraction.

    Extraction has recorded `lines_repaired` since the encoding repair landed,
    but a warehouse built before a given file was scrubbed — or one whose audit
    predates the column — would carry a zero that reads exactly like "this file
    was clean". The markers `ensure_utf8` leaves beside each scrubbed copy are
    the record, so they are the thing to read.

    Loud on every ambiguity: a scrubbed copy with no marker beside it stops the
    run rather than being counted as zero, and a marker no audit row claims
    stops it too. A guessed repair count would be a fabricated data-loss
    figure, which is worse than none.
    """
    ensure_audit_table(con)
    audited = con.execute(
        "SELECT system_id, source_period, source_file FROM raw_file_audit"
    ).fetchall()
    if not audited:
        raise MissingRepairMarker(
            "raw_file_audit is empty — there is nothing to backfill into. Run "
            "the extract stage first."
        )
    by_period: dict[tuple[str, str], list[str]] = {}
    for system_id, period, label in audited:
        by_period.setdefault((system_id, period), []).append(label)

    consumed: set[Path] = set()
    updates: list[tuple[int, str, str, str]] = []
    found = 0
    for (system_id, period), labels in sorted(by_period.items()):
        entry = common.load_manifest(system_id).get("sources", {}).get(period)
        if entry is None:
            raise MissingRepairMarker(
                f"{system_id} {period}: audited but absent from the manifest"
            )
        directory = period_dir(system_id, period, entry)
        for label in labels:
            base = label.split("#", 1)[0]
            if not base.lower().endswith(".csv"):
                # Only CSVs are ever scrubbed: a workbook is read by openpyxl,
                # which decodes XML, not bytes in an unknown code page.
                updates.append((0, system_id, period, label))
                continue
            if directory is None:
                raise MissingRepairMarker(
                    f"{system_id} {period}: the source files are not on disk, so "
                    f"the repair count for {label!r} cannot be read. Restore the "
                    "archive rather than assuming zero."
                )
            scrubbed = directory / SCRUB_DIR_NAME / base
            marker = scrub_marker(directory / base)
            if marker.exists():
                consumed.add(marker.resolve())
                updates.append((int(marker.read_text()), system_id, period, label))
                found += 1
            elif scrubbed.exists():
                raise MissingRepairMarker(
                    f"{system_id} {period} {label}: a UTF-8 repaired copy exists at "
                    f"{scrubbed} with no {marker.name} beside it. How many lines it "
                    "repaired is unknown and will not be guessed — delete the copy "
                    "and re-run the extract stage."
                )
            else:
                updates.append((0, system_id, period, label))

    changed = 0
    for value, system_id, period, label in updates:
        before = con.execute(
            "SELECT coalesce(lines_repaired, -1) FROM raw_file_audit "
            "WHERE system_id = ? AND source_period = ? AND source_file = ?",
            [system_id, period, label],
        ).fetchone()[0]
        if before != value:
            print(f"  {system_id} {period} {label}: {before} -> {value}")
            changed += 1
        con.execute(
            "UPDATE raw_file_audit SET lines_repaired = ? "
            "WHERE system_id = ? AND source_period = ? AND source_file = ?",
            [value, system_id, period, label],
        )

    orphans = sorted(
        p for p in common.DATA_RAW.rglob(f"{SCRUB_DIR_NAME}/*.lines")
        if p.resolve() not in consumed
    )
    if orphans:
        raise MissingRepairMarker(
            f"{len(orphans)} repair marker(s) belong to no audited file, first "
            f"{orphans[0]}. Either the audit is incomplete or the archive holds a "
            "file the manifest does not pin."
        )

    # `kind` is inferable exactly: a station snapshot is a file whose rows
    # landed in raw_stations, which holds a few thousand rows against the
    # fact's 135M. So this reads the small table and never the large one.
    if con.execute("SELECT count(*) FROM information_schema.tables "
                   "WHERE table_name = 'raw_stations'").fetchone()[0]:
        con.execute("""
          UPDATE raw_file_audit a SET kind = CASE WHEN EXISTS (
            SELECT 1 FROM raw_stations s
            WHERE s.system_id = a.system_id AND s.source_period = a.source_period
              AND s.source_file = a.source_file) THEN 'stations' ELSE 'trips' END
          WHERE a.kind IS NULL""")
    record_encoding_metrics(con)
    print(f"backfill: {found} marker(s) read, {changed} audit row(s) changed, "
          f"{len(updates)} file(s) confirmed")
    return 0


def run_sql(con, name: str) -> None:
    con.execute((SQL_DIR / name).read_text(encoding="utf-8"))


def run_model(con) -> None:
    """Build the star schema, with dim_system loaded from the registry."""
    con.execute("SET VARIABLE mappings_dir = ?", [str(common.MAPPINGS_DIR)])
    # Montreal's three key spaces are reconciled before the dimension is built.
    run_sql(con, "35_bridge.sql")
    record(con, "model", "mtl_bridge_entries", "SELECT count(*) FROM mtl_station_bridge")
    run_sql(con, "40_model.sql")

    # dim_system comes from common.SYSTEMS rather than being restated in SQL,
    # so the registry stays the single source of truth for what a system is.
    con.execute("""CREATE OR REPLACE TABLE dim_system (
        system_id VARCHAR PRIMARY KEY, city VARCHAR, system VARCHAR,
        source_page VARCHAR, first_year VARCHAR)""")
    for sid, meta in common.SYSTEMS.items():
        con.execute("INSERT INTO dim_system VALUES (?, ?, ?, ?, ?)",
                    [sid, meta["city"], meta["system"], meta["source_page"],
                     meta["first_year"]])

    record(con, "model", "fact_trips", "SELECT count(*) FROM fact_trips")
    record(con, "model", "dim_station", "SELECT count(*) FROM dim_station")
    record(con, "model", "dim_station_with_coords",
           "SELECT count(*) FROM dim_station WHERE lat IS NOT NULL")
    record(con, "model", "dim_membership", "SELECT count(*) FROM dim_membership")
    record(con, "model", "membership_unmapped",
           "SELECT count(*) FROM dim_membership WHERE membership_group IS NULL")
    record(con, "model", "trips_flagged",
           "SELECT count(*) FROM fact_trips WHERE has_quality_issue")
    record(con, "model", "trips_ebike_known",
           "SELECT count(*) FROM fact_trips WHERE is_ebike IS NOT NULL")


def run_reference(con) -> None:
    """Load the pinned GBFS station feeds.

    This is what makes Montreal's station identity recoverable at all. BIXI has
    used three key spaces across its history — 4-digit codes to 2020, small
    integer `emplacement_pk` in 2021, and station NAMES only from 2022 — and
    the GBFS feed is the one place that carries two of them side by side:
    `short_name` is the old code, `station_id` is the pk. Without it the eras
    are three disjoint sets of stations that happen to describe one network.

    Vancouver's station_id matches the numeric prefix inside its trip-file
    station names; Toronto's matches its published ids directly. Both gain
    coordinates from this, which no trip file supplies.
    """
    con.execute("""CREATE OR REPLACE TABLE gbfs_station (
        system_id VARCHAR, station_id VARCHAR, short_name VARCHAR,
        name VARCHAR, lat DOUBLE, lon DOUBLE, capacity BIGINT)""")
    total = 0
    for system_id in sorted(common.SYSTEMS):
        path = common.reference_path(system_id, "gbfs_station_information")
        if not path.exists():
            print(f"  {system_id}: no GBFS feed on disk — run download.py")
            continue
        payload = json.loads(path.read_text(encoding="utf-8"))
        for st in payload["data"]["stations"]:
            con.execute(
                "INSERT INTO gbfs_station VALUES (?, ?, ?, ?, ?, ?, ?)",
                [system_id, str(st.get("station_id")),
                 str(st["short_name"]) if st.get("short_name") is not None else None,
                 st.get("name"), st.get("lat"), st.get("lon"), st.get("capacity")],
            )
            total += 1
        print(f"  {system_id}: {len(payload['data']['stations']):,} GBFS stations")
    record(con, "reference", "gbfs_stations", "SELECT count(*) FROM gbfs_station")
    load_weather(con)
    return None


def load_weather(con) -> None:
    """ECCC daily climate, one airport station per city.

    A day ECCC does not report stays NULL. It is never zero-filled, never
    forward-filled and never interpolated — 0 °C is a legitimate and common
    value in this dataset, so a zero standing in for a gap would be
    indistinguishable from an observation. The gaps are counted instead, and
    the quality report states them.
    """
    con.execute("""CREATE OR REPLACE TABLE weather_daily (
        system_id VARCHAR, date_key DATE,
        temp_mean_c DOUBLE, temp_max_c DOUBLE, temp_min_c DOUBLE,
        precip_mm DOUBLE, snow_cm DOUBLE, snow_ground_cm DOUBLE)""")
    # Which station each city's weather comes from, carried into the warehouse
    # so the quality report and any published surface can name its source
    # rather than restate it from memory.
    con.execute("""CREATE OR REPLACE TABLE weather_station (
        system_id VARCHAR, station_id VARCHAR, climate_id VARCHAR,
        station_name VARCHAR, lat DOUBLE, lon DOUBLE)""")
    for system_id in sorted(common.SYSTEMS):
        st = (common.load_manifest(system_id) or {}).get("weather_station")
        if st:
            con.execute(
                "INSERT INTO weather_station VALUES (?, ?, ?, ?, ?, ?)",
                [system_id, str(st["station_id"]), str(st["climate_id"]),
                 st["name"], st["lat"], st["lon"]],
            )

    for system_id in sorted(common.SYSTEMS):
        files = sorted(
            (common.DATA_RAW / system_id / "reference").glob("weather_*.csv")
        )
        if not files:
            print(f"  {system_id}: no ECCC files on disk — run download.py")
            continue
        # Read every year at once. ECCC's header carries degree signs and
        # parentheses, so columns are addressed by their exact published names
        # rather than by position — a reordered export must fail loudly, not
        # silently shift precipitation into the temperature column.
        con.execute(
            """
            INSERT INTO weather_daily
            SELECT ?,
                   CAST("Date/Time" AS DATE),
                   TRY_CAST("Mean Temp (°C)"    AS DOUBLE),
                   TRY_CAST("Max Temp (°C)"     AS DOUBLE),
                   TRY_CAST("Min Temp (°C)"     AS DOUBLE),
                   TRY_CAST("Total Precip (mm)" AS DOUBLE),
                   TRY_CAST("Total Snow (cm)"   AS DOUBLE),
                   TRY_CAST("Snow on Grnd (cm)" AS DOUBLE)
            FROM read_csv(?, header = true, all_varchar = true)
            -- ECCC exports a fixed-length calendar year, so the current year
            -- arrives padded with rows for dates that have not happened. Those
            -- carry no observation and are not gaps in the record; keeping
            -- them would report 624 days with no mean temperature where the
            -- table-wide figure is 146.
            --
            -- A row is kept when any of the six STORED measures is present —
            -- not when ECCC measured anything at all. Two van-mobi days
            -- (2025-07-14, 2025-09-17) carry only wind and are dropped here.
            -- Wind is not stored, so nothing is lost, but the distinction is
            -- real and an earlier version of this comment overstated it.
            WHERE coalesce("Mean Temp (°C)", "Max Temp (°C)", "Min Temp (°C)",
                           "Total Precip (mm)", "Total Snow (cm)",
                           "Snow on Grnd (cm)") IS NOT NULL
            """,
            [system_id, [str(p) for p in files]],
        )
        # The CSV names the station it came from. Checking it against the
        # manifest is what makes a mis-set stationID in the bulk URL fail
        # loudly instead of silently storing another city's weather under this
        # system_id — the one error this ingest could otherwise not detect.
        want = con.execute(
            "SELECT climate_id, station_name FROM weather_station WHERE system_id = ?",
            [system_id],
        ).fetchone()
        if want:
            found = con.execute(
                'SELECT DISTINCT "Climate ID", "Station Name" '
                "FROM read_csv(?, header = true, all_varchar = true)",
                [[str(p) for p in files]],
            ).fetchall()
            unexpected = [f for f in found if f[0] != want[0]]
            if unexpected:
                raise SystemExit(
                    f"{system_id}: ECCC files carry climate ID(s) {unexpected} "
                    f"but the manifest declares {want[0]} ({want[1]}). The "
                    "station in the download URL does not match the one this "
                    "system is supposed to use."
                )

        obs, gaps, first, last = con.execute(
            """SELECT count(*), count(*) FILTER (temp_mean_c IS NULL),
                      min(date_key), max(date_key)
               FROM weather_daily WHERE system_id = ?""",
            [system_id],
        ).fetchone()
        print(
            f"  {system_id}: {obs:,} days {first}..{last}, "
            f"{gaps:,} with no mean temperature"
        )

    record(con, "reference", "weather_days", "SELECT count(*) FROM weather_daily")
    record(
        con, "reference", "weather_days_missing_temp",
        "SELECT count(*) FROM weather_daily WHERE temp_mean_c IS NULL",
    )


def run_clean(con) -> None:
    """Resolve per-file date order, clean, then account for every dropped row.

    The funnel must close: landed = kept + dropped-by-reason. If it does not,
    something is being lost silently, which is the failure this project's
    quality report exists to make impossible.
    """
    # Must precede clean: the parser reads each file's proven date order.
    run_sql(con, "15_dates.sql")
    run_sql(con, "17_timezones.sql")
    for sys_id, f, basis, trough, n in con.execute(
        "SELECT system_id, source_file, tz_basis, trough_hour, rows_parsed "
        "FROM file_timezone WHERE tz_basis <> 'local' ORDER BY 1, 2"
    ).fetchall():
        print(f"  timezone {basis.upper()} for {sys_id} {f} "
              f"(trough {trough:02d}:00, {n:,} rows)", flush=True)
    record(con, "clean", "files_utc",
           "SELECT count(*) FROM file_timezone WHERE tz_basis = 'utc'")
    record(con, "clean", "files_tz_unknown",
           "SELECT count(*) FROM file_timezone WHERE tz_basis = 'unknown'")
    odd = con.execute(
        "SELECT system_id, source_file, date_order FROM file_date_order "
        "WHERE date_order IN ('conflict', 'ambiguous')"
    ).fetchall()
    for sys_id, f, ord_ in odd:
        print(f"  date order {ord_.upper()} for {sys_id} {f} — parsed month-first", flush=True)
    record(con, "clean", "files_day_first",
           "SELECT count(*) FROM file_date_order WHERE date_order = 'day'")
    record(con, "clean", "files_ambiguous_date_order",
           "SELECT count(*) FROM file_date_order WHERE date_order IN ('conflict','ambiguous')")
    run_sql(con, "20_clean.sql")
    run_sql(con, "25_localise.sql")
    record(con, "clean", "rows_kept", "SELECT count(*) FROM clean_trips")
    # These MUST use the same context-aware parser clean_trips uses. When they
    # did not, the funnel produced a negative duplicate count — which is exactly
    # what the derived-and-checked design is for.
    record(
        con, "clean", "rows_dropped_no_departure_time",
        """SELECT count(*) FROM raw_trips r
           LEFT JOIN file_date_order o
             ON o.system_id = r.system_id AND o.source_file = r.source_file
           WHERE coalesce(parse_ts_ord(r.departure_raw, o.date_order),
                          epoch_ms(try_cast(r.departure_ms AS BIGINT))) IS NULL""",
    )
    # Normalised, not merely trimmed — the same test 20_clean.sql applies. The
    # trimmed form counted the literal token 'NULL' as a station name, so five
    # Toronto rows were neither dropped nor countable under any reason.
    record(
        con, "clean", "rows_dropped_no_departure_station",
        """SELECT count(*) FROM raw_trips r
           LEFT JOIN file_date_order o
             ON o.system_id = r.system_id AND o.source_file = r.source_file
           WHERE coalesce(parse_ts_ord(r.departure_raw, o.date_order),
                          epoch_ms(try_cast(r.departure_ms AS BIGINT))) IS NOT NULL
             AND norm_key(r.departure_station_key) IS NULL
             AND norm_name(r.departure_station_name) IS NULL""",
    )
    # Duplicates are COUNTED, not inferred. This was previously landed minus
    # kept minus the other reasons, which closes the funnel by construction:
    # the residual below could not be anything but zero, so it tested nothing.
    # A row lands here only if it would otherwise have been kept — it has a
    # departure time and a departure station and lost the dedupe — so the three
    # reasons stay disjoint and the sum is a real claim.
    record(
        con, "clean", "rows_dropped_duplicates",
        """SELECT count(*) FROM raw_trips r
           LEFT JOIN file_date_order o
             ON o.system_id = r.system_id AND o.source_file = r.source_file
           WHERE r.system_id = 'van-mobi'
             AND coalesce(parse_ts_ord(r.departure_raw, o.date_order),
                          epoch_ms(try_cast(r.departure_ms AS BIGINT))) IS NOT NULL
             AND (norm_key(r.departure_station_key) IS NOT NULL
                  OR norm_name(r.departure_station_name) IS NOT NULL)
             AND r.rowid NOT IN (SELECT keep_rid FROM van_dedupe_keep)""",
    )
    record(con, "clean", "rows_unterminated",
           "SELECT count(*) FROM clean_trips WHERE return_ts IS NULL")
    record(con, "clean", "station_rows", "SELECT count(*) FROM clean_stations")
    assert_funnel_closes(con)


DROP_REASONS = (
    "rows_dropped_no_departure_time",
    "rows_dropped_no_departure_station",
    "rows_dropped_duplicates",
)


def funnel(con) -> tuple[int, int, list[tuple[str, int]], int]:
    """(landed, kept, [(reason, rows)], residual) — every term counted alone."""
    metrics = dict(
        ((s, m), v) for s, m, v in
        con.execute("SELECT stage, metric, value FROM etl_metrics").fetchall()
    )
    landed = metrics.get(("extract", "rows_landed"), 0)
    kept = metrics.get(("clean", "rows_kept"), 0)
    named = [(r, metrics.get(("clean", r), 0)) for r in DROP_REASONS]
    return landed, kept, named, landed - kept - sum(n for _, n in named)


def assert_funnel_closes(con) -> None:
    landed, kept, named, residual = funnel(con)
    for reason, n in named:
        print(f"clean.{reason} = {n:,}")
    if residual:
        raise SystemExit(
            f"funnel does not close: {landed:,} landed, {kept:,} kept, "
            + ", ".join(f"{n:,} {r}" for r, n in named)
            + f" — residual {residual:,}. Every drop is counted at the stage that "
            "makes it, so a non-zero residual means rows are leaving the pipeline "
            "under no reason at all, or one reason is counting another's rows."
        )
    print(f"funnel closes: {landed:,} landed = {kept:,} kept + "
          f"{sum(n for _, n in named):,} dropped")


def record(con, stage: str, metric: str, query: str) -> None:
    con.execute("DELETE FROM etl_metrics WHERE stage = ? AND metric = ?", [stage, metric])
    value = con.execute(query).fetchone()[0]
    con.execute("INSERT INTO etl_metrics VALUES (?, ?, ?)", [stage, metric, value])
    print(f"{stage}.{metric} = {value:,}")


def connect(db_path: Path) -> duckdb.DuckDBPyConnection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(str(db_path))
    con.execute("INSTALL excel; LOAD excel;")
    con.execute("INSTALL icu; LOAD icu;")  # timezone conversion in 20_clean.sql
    # 135M rows on a 16 GB machine: bound memory explicitly and give the spill
    # a home on the data volume. Left to its own devices DuckDB will take ~80%
    # of RAM and then thrash.
    con.execute("SET preserve_insertion_order = false;")
    con.execute("SET memory_limit = '10GB';")
    con.execute("SET threads = 8;")
    (common.DATA_WAREHOUSE / "tmp").mkdir(parents=True, exist_ok=True)
    con.execute(f"SET temp_directory = '{common.DATA_WAREHOUSE / 'tmp'}';")
    return con


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--stage", choices=STAGES + ("all",), default="all")
    parser.add_argument("--system", choices=sorted(common.SYSTEMS), action="append")
    parser.add_argument("--limit", type=int)
    parser.add_argument("--db", type=Path, default=WAREHOUSE)
    parser.add_argument(
        "--backfill-encoding-repairs", action="store_true",
        help="read the UTF-8 repair markers on disk into raw_file_audit and "
             "exit, without re-running extraction",
    )
    args = parser.parse_args()

    systems = args.system or sorted(common.SYSTEMS)
    con = connect(args.db)
    if args.backfill_encoding_repairs:
        try:
            return backfill_encoding_repairs(con)
        except MissingRepairMarker as exc:
            print(f"\nABORT: {exc}", file=sys.stderr)
            return 1
        finally:
            con.close()
    stages = STAGES if args.stage == "all" else (args.stage,)
    try:
        for stage in stages:
            print(f"\n=== {stage}")
            if stage == "extract":
                run_extract(con, systems, args.limit)
            elif stage == "reference":
                run_reference(con)
            elif stage == "clean":
                run_clean(con)
            elif stage == "model":
                run_model(con)
            else:
                run_sql(con, SQL_FILES[stage])
    except (UnknownColumns, UnknownMember, ReconciliationFailed,
            FileNotFoundError) as exc:
        print(f"\nABORT: {exc}", file=sys.stderr)
        return 1
    finally:
        con.close()
    print(f"\nwarehouse: {args.db}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
